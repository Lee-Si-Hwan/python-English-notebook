# python-English-notebook

from tkinter import *
import random
import easygui
from openpyxl import load_workbook

#데이터를 불러옴
filename = "data.xlsx"
data = load_workbook(filename)
sheet = data.worksheets[0]
sheet_rows = sheet.max_row
sheet_cols = sheet.max_column

root = Tk()


#모드 변수
mn = ''

#처음에 모드 설정 창
def menu_mode():
    global mn
    mn = easygui.buttonbox("choice to want to mode. I have "+str(sheet_rows)+" words", "menu", ("set word", 'word mode', 'meaning mode', 'finding word')) #메뉴 설정창
    menu()
    

def menu():
    if mn == 'set word':
        title.config(text = mn)
        b1.config(text = 'save')
        ask.config(text = '1.영단어를 입력하시오')
        lb1.config(text = '2.뜻을 한글로 입력하시오')

    elif mn == 'word mode':
        title.config(text = mn)
        b1.config(text = 'pass')
        setting()


    elif mn == 'meaning mode':
        title.config(text = mn)
        b1.config(text = 'pass')
        setting()

    else:
        find()



sheetlist = []
asks = ''
answers = ''
k = 0
l = 0
x = 0


def setting():
    for i in range(len(sheetlist)):
        del sheetlist[0]
        
    for i in range(1,sheet_rows+1):
        sheetlist.append(i)
    print(sheetlist)
    rand()        
        
        
        
def nextword():
    if mn == 'set word':
        print('save')
    if mn == 'word mode':
        print()
        
    
def All(event):
    if mn == "set word":
        if ask['text'] == '1.영단어를 입력하시오':
            ask.config(text = answer.get())
            lb1.config(text = '')
        else:
            means = lb1['text']
            if means == '':
                means = answer.get()
            else:
                means += ','
                means += answer.get()
            lb1.config(text = means)


    elif mn == 'word mode':
        wordQ()
    elif mn == 'meaning mode':
        meaningQ()
    else:
        search()


def wordQ():                             #뜻 맞추기
    if mn == 'set word':
        print(10)


    else:
        global l
        global x
        l = 2

        r = 0

        while True:
            if answer.get() == answers:
                lb1.config(text = '정답')
                break
    
            else:
                while sheet.cell(x,l).value != '.':
                    l += 1

                    if answer.get() == sheet.cell(x,l).value:
                        lb1.config(text = '정답')
                        r = 1
                        break
                    
                if r == 1:
                    print('bingo')

                else:
                    lb1.config(text = '땡')
                    l = 2
                break
            
        if lb1["text"] == '정답':
            rand()
        
def meaningQ():                  #단어 맞추기
    global l
    global x

    if answer.get() == answers:
        lb1.config(text = '정답')
        rand()
    else:
        lb1.config(text = '땡!')


def find():
    title.config(text = 'finding word')
    ask.config(text = '찾고 싶은 단어를 입력하세요')
    b1.config(text = 'search')
    


def next():
    global mn
    if mn == 'set word':
        append()
    if mn == 'finding word':
        search()
    else:
        rand()
    
#엑셀 파일에 정보 저장하는 함수
def append():
    append_list = []
    append_list.append(ask['text'])
    means = lb1['text']
    meansnum = len(means)
    value = ''
    for i in range(len(means)):
        if means[i] == ',' or i == meansnum - 1: #단어와 ','구별
            if i == meansnum - 1:
                value += means[i]
            append_list.append(value)
            value = ''                        #리셋
        else:
            value += means[i]
    print(append_list)
    append_list.append(".")
    sheet.append(append_list)                 #정보 입력
    data.save("data.xlsx")                    #저장
    lb1.config(text = '저장되었습니다. 저장한 단어로 문제를 풀려면 프로그램을 다시 실행하세요.')
    ask.config(text = '1.영단어를 입력하시오')

def search():
    searchlist = []
    mainlist = []
    printlist = []
    
    value = answer.get()
    for row in range(1, sheet_rows+1):
        for col in range(1, sheet_cols+1):
            if value == sheet.cell(row,col).value:
                searchlist.append(row)
                
    if len(searchlist) == 0:
        print('없음')
    else:
        for i in searchlist:
            col = 1
            while sheet.cell(i,col).value != '.':
                printlist.append(sheet.cell(i,col).value + ',')
                col += 1
            mainlist.append(str(printlist)+'\n')
            printlist.clear()
            
    if len(searchlist) == 0:
        easygui.msgbox('None','text')
    else:
        easygui.msgbox(mainlist,"show")
                

#문제 푸는 함수
def rand():
    global asks
    global answers
    global x
    
    means = []
    msglist = []
    askword = ''

    line.config(text = '----------------------------------'+str(len(sheetlist))+'개 남음')

    if len(sheetlist) == 0:
        ask.config(text = '끝났습니다! back을 누르고 모드를 선택하세요.')

    else:
        if lb1["text"] == '땡!':
                row = x
                col = 1
                value = sheet.cell(row,col).value
                while value != '.':
                    msglist.append(value + ',')
                    col += 1
                    value = sheet.cell(row,col).value
                easygui.msgbox(msglist, "test")
        lb1.config(text = '')
        y = random.sample(sheetlist, 1)
        x = y[0]
        #한 번 출제된 문제 다신 안나오게 하기
        sheetlist.remove(x)
        
        if mn == 'word mode':
            k = 1
            l = 2
            
            asks = sheet.cell(x,k).value
            answers = sheet.cell(x,l).value
        else:
            l = 2
            while sheet.cell(x,l).value != '.':
                means.append(l)
                l += 1

            for i in range(2,len(means)+2):
                askword += sheet.cell(x,i).value
                askword += ','

            asks = askword
            answers = sheet.cell(x,1).value

        ask.config(text = asks)


#단어 맞추는 메인 창 설정

title = Label(root, text = mn)

line = Label(root, text = '----------------------------------')
ask = Label(root, text = asks)


answer = Entry(root)
answer.bind("<Return>", All)


lb1 = Label(root, text = '')

b1 = Button(root, text = 'pass', command = next)

b2 = Button(root, text = 'back', command = menu_mode)

menu_mode()

title.pack()
line.pack()
ask.pack()
answer.pack()
lb1.pack()
b1.pack()
b2.pack()


root.mainloop()

##########################
#단어 추가는 되는데 그건 파이썬.idle로만 되고 비주얼코드는 자꾸 easygui가 문제있다면서 안됨(visual studio code에서만)
#다음에 제발 고쳐줘

